from __future__ import annotations

import math
import os
import re
import shutil
import subprocess
import tempfile
from copy import copy
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .errors import OutputExistsError, TemplateError, ValidationError
from .models import ExportRequest, ExpenseLine

DETAIL_START = 11
DETAIL_END = 30
ROWS_PER_PAGE = 20
REQUIRED_CELLS = ("A3", "A5", "C7", "A11", "F11", "G11", "F31", "E44")


@dataclass(frozen=True)
class ExportResult:
    path: Path
    detail_count: int
    page_count: int
    total: int
    page_subtotals: tuple[int, ...]
    line_ids: tuple[str, ...]


class TravelExpenseExporter:
    def __init__(self, soffice: str = "soffice") -> None:
        self.soffice = soffice

    def export(self, template: Path, request: ExportRequest, output: Path | None = None) -> ExportResult:
        template = template.resolve()
        if not template.is_file():
            raise TemplateError("出張旅費精算書の原本ファイルが見つかりません。")
        lines = self._select_lines(request)
        self._validate_request(request, lines)
        page_count = max(1, math.ceil(len(lines) / ROWS_PER_PAGE))
        output = (output or request.output_dir / self.default_filename(request)).resolve()
        if output.exists():
            raise OutputExistsError(f"出力先ファイルは既に存在します: {output}")
        output.parent.mkdir(parents=True, exist_ok=True)

        with tempfile.TemporaryDirectory(prefix="travel-expense-") as work:
            source = self._working_xlsx(template, Path(work))
            workbook = load_workbook(source)
            self._validate_template(workbook, request.template_sheet)
            base = workbook[request.template_sheet]
            sheets = self._prepare_sheets(workbook, base, page_count, request.include_sample_sheet)
            subtotals = []
            for index, sheet in enumerate(sheets):
                chunk = lines[index * ROWS_PER_PAGE : (index + 1) * ROWS_PER_PAGE]
                self._write_sheet(sheet, request, chunk, index + 1, page_count)
                subtotals.append(sum(line.amount for line in chunk))

            temporary_output = Path(work) / "result.xlsx"
            workbook.save(temporary_output)
            self._verify_saved(temporary_output, sheets, subtotals)
            try:
                os.replace(temporary_output, output)
            except OSError:
                shutil.copy2(temporary_output, output)

        return ExportResult(
            path=output,
            detail_count=len(lines),
            page_count=page_count,
            total=sum(line.amount for line in lines),
            page_subtotals=tuple(subtotals),
            line_ids=tuple(line.id for line in lines),
        )

    def default_filename(self, request: ExportRequest) -> str:
        name = re.sub(r'[\\/:*?"<>|\s]+', "", request.employee_name)
        suffix = f"_修正版{request.revision}" if request.revision else ""
        return f"{request.year}年{request.month:02d}月_出張旅費代精算書_{name}{suffix}.xlsx"

    def _select_lines(self, request: ExportRequest) -> list[ExpenseLine]:
        selected = [line for line in request.lines if line.is_exportable(request.year, request.month)]
        return sorted(
            selected,
            key=lambda line: (
                line.travel_date,
                line.start_time is None,
                line.start_time or __import__("datetime").time.max,
                line.route_order,
                line.created_order,
            ),
        )

    def _validate_request(self, request: ExportRequest, lines: list[ExpenseLine]) -> None:
        missing = []
        if not request.department.strip(): missing.append("所属")
        if not request.employee_name.strip(): missing.append("氏名")
        if request.commuter_pass is None: missing.append("定期区間または定期券なし")
        if not request.submission_date: missing.append("提出日")
        if not lines:
            raise ValidationError("出力対象となる確認済み旅費がありません。")
        for number, line in enumerate(lines, 1):
            if not line.destination.strip(): missing.append(f"{number}行目の目的地")
            if not line.section_text().strip(): missing.append(f"{number}行目の有料区間")
            if not line.reason.strip(): missing.append(f"{number}行目の移動理由")
            if line.transit_parts and sum(p.amount for p in line.transit_parts) != line.amount:
                raise ValidationError(f"{number}行目の交通手段別内訳と申請額が一致しません。")
        if missing:
            raise ValidationError("必須項目が未登録です: " + "、".join(missing))

    def _working_xlsx(self, template: Path, work: Path) -> Path:
        if template.suffix.lower() == ".xlsx":
            destination = work / "template.xlsx"
            shutil.copy2(template, destination)
            return destination
        if template.suffix.lower() != ".ods":
            raise TemplateError("原本は.xlsxまたは.odsを指定してください。")
        copied = work / template.name
        shutil.copy2(template, copied)
        try:
            completed = subprocess.run(
                [self.soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(work), str(copied)],
                capture_output=True, text=True, timeout=120, check=False,
            )
        except (OSError, subprocess.TimeoutExpired) as exc:
            raise TemplateError("ODSをXLSXへ変換できません。LibreOfficeの設定を確認してください。") from exc
        converted = work / f"{template.stem}.xlsx"
        if completed.returncode or not converted.exists():
            raise TemplateError("ODSをXLSXへ変換できません: " + (completed.stderr or completed.stdout).strip())
        return converted

    def _validate_template(self, workbook, sheet_name: str) -> None:
        if sheet_name not in workbook.sheetnames:
            raise TemplateError(f"原本シート「{sheet_name}」が見つかりません。")
        sheet = workbook[sheet_name]
        for cell in REQUIRED_CELLS:
            sheet[cell]  # 座標が有効であることを確認
        merged = {str(rng) for rng in sheet.merged_cells.ranges}
        for row in range(DETAIL_START, DETAIL_END + 1):
            if f"D{row}:E{row}" not in merged or f"G{row}:H{row}" not in merged:
                raise TemplateError("原本の明細結合セル構造が変更されています。")
        formula = sheet["F31"].value
        if not isinstance(formula, str) or not formula.startswith("="):
            raise TemplateError("F31に合計数式がありません。")

    def _prepare_sheets(self, workbook, base: Worksheet, count: int, include_sample: bool) -> list[Worksheet]:
        sheets = []
        for index in range(count):
            sheet = base if index == 0 else workbook.copy_worksheet(base)
            sheet.title = f"出張旅費精算_{index + 1}"
            sheets.append(sheet)
        for sheet in list(workbook.worksheets):
            if sheet not in sheets and (not include_sample or sheet.title != "【見本】出張旅費精算"):
                workbook.remove(sheet)
        return sheets

    def _write_sheet(self, sheet: Worksheet, request: ExportRequest, lines: list[ExpenseLine], page: int, pages: int) -> None:
        sheet["C1"] = request.department
        sheet["G1"] = f"（ {page} ）枚目/（ {pages} ）枚中"
        sheet["A3"] = f"{request.month}月度 出張旅費代精算書（電車・バス用）"
        sheet["A5"] = f"氏名　{request.employee_name}"
        sheet["C7"] = request.commuter_pass or "定期券なし"
        if request.pass_changes:
            sheet["A8"] = "\n".join(
                f"※月度途中で区間変更⇒（{change.effective_date.month}月{change.effective_date.day}日より {change.new_section}区間に変更）"
                for change in request.pass_changes
            )
        sheet["E44"] = f"{request.submission_date.year}年 {request.submission_date.month}月 {request.submission_date.day}日"
        for row in range(DETAIL_START, DETAIL_END + 1):
            for column in ("A", "B", "C", "D", "F", "G"):
                sheet[f"{column}{row}"] = None
        for row, line in enumerate(lines, DETAIL_START):
            sheet[f"A{row}"] = line.travel_date.month
            sheet[f"B{row}"] = line.travel_date.day
            sheet[f"C{row}"] = line.destination
            sheet[f"D{row}"] = line.section_text()
            sheet[f"F{row}"] = line.amount
            sheet[f"G{row}"] = line.reason
        sheet["F31"] = "=SUM(F11:F30)"
        sheet.print_area = "A1:H44"
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1

    def _verify_saved(self, path: Path, sheets: list[Worksheet], subtotals: list[int]) -> None:
        check = load_workbook(path, data_only=False)
        names = [sheet.title for sheet in sheets]
        if any(name not in check.sheetnames for name in names):
            raise TemplateError("保存後に提出用シートが見つかりません。")
        for index, name in enumerate(names):
            sheet = check[name]
            if sheet["F31"].value != "=SUM(F11:F30)":
                raise TemplateError("保存後にF31の合計数式が破損しました。")
            merged = {str(rng) for rng in sheet.merged_cells.ranges}
            if any(f"D{r}:E{r}" not in merged or f"G{r}:H{r}" not in merged for r in range(11, 31)):
                raise TemplateError("保存後に明細の結合セルが破損しました。")
            actual = sum((sheet[f"F{row}"].value or 0) for row in range(11, 31))
            if actual != subtotals[index]:
                raise TemplateError("保存後のページ小計がアプリ内小計と一致しません。")
