from __future__ import annotations

import tempfile
import unittest
from datetime import date, time
from pathlib import Path

from openpyxl import load_workbook

from tests.conftest import make_template
from travel_expense import ExportRequest, ExpenseLine, TransitPart, TravelExpenseExporter
from travel_expense.errors import OutputExistsError, ValidationError


def line(n: int, amount: int = 100, **kwargs) -> ExpenseLine:
    return ExpenseLine(
        id=str(n), travel_date=kwargs.pop("travel_date", date(2026, 7, (n % 28) + 1)),
        destination=f"目的地{n}", paid_section=f"駅{n}→駅{n+1}", amount=amount,
        reason="学校訪問", created_order=n, **kwargs,
    )


def request(lines, tmp_path, **kwargs):
    return ExportRequest(
        year=2026, month=7, department="さいたま本部", employee_name="藤野 恒児",
        commuter_pass="ふじみ野駅 ～ 池袋駅", submission_date=date(2026, 8, 3),
        lines=tuple(lines), output_dir=tmp_path, **kwargs,
    )


class ExporterTests(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory()
        self.tmp_path = Path(self.temp.name)
        self.template = self.tmp_path / "template.xlsx"
        make_template(self.template)

    def tearDown(self):
        self.temp.cleanup()

    def test_one_line_and_formula_and_template_unchanged(self):
        before = self.template.read_bytes()
        result = TravelExpenseExporter().export(
            self.template, request([line(14, 406, travel_date=date(2026, 7, 15))], self.tmp_path)
        )
        ws = load_workbook(result.path, data_only=False)["出張旅費精算_1"]
        self.assertEqual([ws[c].value for c in ("A11", "B11", "C11", "D11", "F11", "G11")],
                         [7, 15, "目的地14", "駅14→駅15", 406, "学校訪問"])
        self.assertEqual(ws["F31"].value, "=SUM(F11:F30)")
        self.assertEqual(self.template.read_bytes(), before)

    def test_page_splits(self):
        for count, pages, sizes in [(20, 1, [20]), (21, 2, [20, 1]), (45, 3, [20, 20, 5])]:
            with self.subTest(count=count):
                case_dir = self.tmp_path / str(count)
                case_dir.mkdir()
                lines = [line(n, travel_date=date(2026, 7, 1 + n // 3), start_time=time(n % 3 + 8)) for n in range(count)]
                result = TravelExpenseExporter().export(self.template, request(lines, case_dir))
                wb = load_workbook(result.path)
                self.assertEqual(len(wb.worksheets), pages)
                self.assertEqual(result.page_subtotals, tuple(size * 100 for size in sizes))
                self.assertEqual([sum(ws[f"F{r}"].value is not None for r in range(11, 31)) for ws in wb.worksheets], sizes)

    def test_filter_sort_and_transit_breakdown(self):
        parts = (TransitPart("電車", "大宮", "浦和", 178), TransitPart("バス", "浦和駅", "原山3丁目", 220))
        lines = [line(1, 0), line(2, excluded=True), line(3, submitted=True),
                 line(4, 398, travel_date=date(2026, 7, 15), transit_parts=parts),
                 line(5, 100, travel_date=date(2026, 7, 14))]
        result = TravelExpenseExporter().export(self.template, request(lines, self.tmp_path))
        ws = load_workbook(result.path).active
        self.assertEqual(result.detail_count, 2)
        self.assertEqual(ws["B11"].value, 14)
        self.assertEqual(ws["D12"].value, "（電車）大宮→浦和 178円\n（バス）浦和駅→原山3丁目 220円")
        self.assertEqual(ws["F12"].value, 398)

    def test_invalid_breakdown_and_no_overwrite(self):
        bad = line(1, 398, transit_parts=(TransitPart("電車", "A", "B", 100),))
        with self.assertRaises(ValidationError):
            TravelExpenseExporter().export(self.template, request([bad], self.tmp_path))
        good_request = request([line(2)], self.tmp_path)
        TravelExpenseExporter().export(self.template, good_request)
        with self.assertRaises(OutputExistsError):
            TravelExpenseExporter().export(self.template, good_request)

    def test_optional_sample_sheet_and_empty_selection(self):
        result = TravelExpenseExporter().export(
            self.template, request([line(1)], self.tmp_path, include_sample_sheet=True)
        )
        self.assertEqual(load_workbook(result.path).sheetnames,
                         ["出張旅費精算_1", "【見本】出張旅費精算"])
        empty_dir = self.tmp_path / "empty"
        with self.assertRaises(ValidationError):
            TravelExpenseExporter().export(self.template, request([line(2, 0)], empty_dir))


if __name__ == "__main__":
    unittest.main()
