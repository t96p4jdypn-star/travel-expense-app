from __future__ import annotations

from openpyxl import Workbook


def make_template(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "【原本】出張旅費精算"
    ws.merge_cells("C1:D1")
    ws.merge_cells("G1:H1")
    ws.merge_cells("A3:H3")
    ws.merge_cells("C7:F7")
    for row in range(11, 31):
        ws.merge_cells(f"D{row}:E{row}")
        ws.merge_cells(f"G{row}:H{row}")
    ws["A3"] = "6月度 出張旅費代精算書"
    ws["A5"] = "氏名"
    ws["C7"] = "駅 ～ 駅"
    ws["F31"] = "=SUM(F11:F30)"
    ws["E44"] = "2026年 月 日"
    sample = wb.create_sheet("【見本】出張旅費精算")
    sample["A1"] = "見本"
    wb.save(path)
